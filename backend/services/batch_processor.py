"""
MetaLex Batch Product Scanning & Consolidated Reporting Service
Accepts ZIP archives of packaged product photos, processes in parallel,
tracks real-time progress, and produces downloadable Excel summary reports.
"""

import os
import io
import time
import uuid
import zipfile
import asyncio
import logging
from pathlib import Path
from typing import Dict, Any, List, Optional
from concurrent.futures import ThreadPoolExecutor

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from backend.database import SessionLocal
from backend.models.db_models import Inspection, ExtractedField, Violation, User
from backend.services.ocr_pipeline import process_single_image
from backend.services.barcode_detector import verify_against_gs1
from rules.rule_engine import get_rule_engine
from backend.services.commodity_detector import detect_commodity_category

logger = logging.getLogger("metalex.batch")

# In-memory batch job tracking
BATCH_JOBS: Dict[str, Dict[str, Any]] = {}

MAX_BATCH_FILES = 50
MAX_BATCH_SIZE_BYTES = 100 * 1024 * 1024  # 100MB
ALLOWED_IMG_EXTS = {".jpg", ".jpeg", ".png", ".bmp", ".webp", ".tiff"}

_executor = ThreadPoolExecutor(max_workers=4)


def get_batch_job(batch_id: str) -> Optional[Dict[str, Any]]:
    return BATCH_JOBS.get(batch_id)


def list_batch_jobs() -> List[Dict[str, Any]]:
    return sorted(BATCH_JOBS.values(), key=lambda j: j.get("created_at", 0), reverse=True)


def init_batch_job(batch_id: str, total_files: int, zip_filename: str) -> Dict[str, Any]:
    job = {
        "batch_id": batch_id,
        "filename": zip_filename,
        "status": "PROCESSING",  # PROCESSING, COMPLETED, FAILED
        "progress_pct": 0,
        "processed_count": 0,
        "total_count": total_files,
        "compliant_count": 0,
        "non_compliant_count": 0,
        "needs_review_count": 0,
        "inspections": [],
        "created_at": time.time(),
        "completed_at": None,
        "duration_seconds": 0,
        "error": None,
    }
    BATCH_JOBS[batch_id] = job
    return job


def process_single_batch_item(
    image_bytes: bytes,
    filename: str,
    inspection_id: str,
    upload_dir: Path,
    annotated_dir: Path,
    user_id: Optional[str] = None
) -> Dict[str, Any]:
    """Process a single product image in a standalone database session."""
    db = SessionLocal()
    try:
        ext = Path(filename).suffix.lower() or ".jpg"
        img_filename = f"{inspection_id}_0{ext}"
        img_path = upload_dir / img_filename
        
        with open(img_path, "wb") as f:
            f.write(image_bytes)

        # Run OCR + Vision Pipeline
        ocr_res = process_single_image(str(img_path))
        fields = ocr_res.get("fields", {})
        barcodes = ocr_res.get("barcodes", [])

        # Run Barcode verification
        barcode_summary = None
        barcode_info_for_engine = {"detected": False}
        if barcodes:
            primary_bc = barcodes[0]
            bc_data = primary_bc.get("data")
            if bc_data:
                bc_verify = verify_against_gs1(bc_data, fields)
                barcode_summary = bc_verify
                barcode_info_for_engine = {
                    "detected": True,
                    "barcode": bc_data,
                    "gs1_found": bc_verify.get("gs1_found", False),
                    "mismatches": bc_verify.get("mismatches", []),
                    "gs1_product_name": bc_verify.get("gs1_product_name"),
                    "gs1_manufacturer": bc_verify.get("gs1_manufacturer"),
                }

        fields_for_engine = dict(fields)
        fields_for_engine["barcode"] = {
            "value": barcodes[0]["data"] if barcodes else None,
            "confidence": 1.0 if barcodes else 0.0,
            "barcode_info": barcode_info_for_engine,
            "bounding_box": barcodes[0].get("bbox") if barcodes else None
        }

        # Run Legal Metrology Rule Engine
        engine = get_rule_engine()
        validation = engine.validate_all(fields_for_engine)

        product_name = fields.get("product_name", {}).get("value") or Path(filename).stem.replace("_", " ").title()

        # Save annotated image
        ann_img = ocr_res.get("annotated_image")
        annotated_path = None
        if ann_img is not None:
            import cv2
            ann_filename = f"{inspection_id}_0_annotated.jpg"
            ann_full_path = annotated_dir / ann_filename
            cv2.imwrite(str(ann_full_path), ann_img)
            annotated_path = str(ann_full_path)

        # Commodity category
        all_text = " ".join([str(f.get("source_text", "")) for f in fields.values()])
        commodity_res = detect_commodity_category(ocr_text=all_text, product_name=product_name)

        # DB Record
        inspection = Inspection(
            inspection_id=inspection_id,
            product_name=product_name,
            image_path=str(img_path),
            annotated_image_path=annotated_path,
            overall_status=validation["overall_status"],
            compliance_score=validation["compliance_score"],
            severity_score=validation.get("severity_score", 0.0),
            risk_level=validation.get("risk_level", "low"),
            barcode_data=barcode_summary,
            total_fields=validation["total_checks"],
            passed_fields=validation["passed"],
            failed_fields=validation["failed"],
            review_fields=validation["needs_review"],
            is_demo=False,
            image_quality_score=ocr_res.get("quality", {}).get("overall_score", 0.9),
            ocr_engine=ocr_res.get("ocr_engine", "easyocr_multipass_v3"),
            commodity_category=commodity_res.get("category"),
            user_id=user_id,
        )
        db.add(inspection)
        db.flush()

        # Save fields
        for fname, fdata in fields.items():
            ef = ExtractedField(
                inspection_id=inspection.id,
                field_name=fname,
                field_label=fname.replace("_", " ").title(),
                detected_value=fdata.get("value"),
                normalized_value=fdata.get("normalized_value"),
                confidence=fdata.get("confidence", 0.0),
                status="PASS" if fdata.get("value") else "FAIL",
                bounding_box=fdata.get("bounding_box"),
                font_size_mm=fdata.get("font_size_mm"),
                min_font_size_mm=fdata.get("min_font_size_mm"),
                source_text=fdata.get("source_text", ""),
                extraction_method=fdata.get("extraction_method", "ocr"),
            )
            db.add(ef)

        # Save violations
        for v in validation.get("violations", []):
            viol = Violation(
                inspection_id=inspection.id,
                rule_id=v["rule_id"],
                field=v["field"],
                severity=v.get("severity", "high"),
                severity_points=v.get("severity_points", 5),
                title=v["rule_title"],
                detected_value=v.get("detected_value"),
                expected_requirement=v.get("expected_requirement"),
                reason=v.get("reason", ""),
                confidence=v.get("confidence"),
                evidence_type=v.get("evidence_type", "image"),
            )
            db.add(viol)

        db.commit()

        return {
            "id": inspection.id,
            "inspection_id": inspection.inspection_id,
            "product_name": inspection.product_name,
            "filename": filename,
            "status": inspection.overall_status,
            "compliance_score": inspection.compliance_score,
            "severity_score": inspection.severity_score,
            "risk_level": inspection.risk_level,
            "violations_count": len(validation.get("violations", [])),
            "mrp": fields.get("mrp", {}).get("value"),
            "net_quantity": fields.get("net_quantity", {}).get("value"),
            "barcode": barcodes[0]["data"] if barcodes else None,
            "success": True,
        }
    except Exception as e:
        logger.error(f"Error processing batch item {filename}: {e}")
        db.rollback()
        return {
            "inspection_id": inspection_id,
            "filename": filename,
            "status": "ERROR",
            "compliance_score": 0,
            "severity_score": 100,
            "risk_level": "critical",
            "error": str(e),
            "success": False,
        }
    finally:
        db.close()


async def process_batch_zip(
    zip_bytes: bytes,
    zip_filename: str,
    upload_dir: Path,
    annotated_dir: Path,
    user_id: Optional[str] = None
) -> str:
    """Unpack ZIP archive and process all valid product images in parallel."""
    batch_id = f"BATCH-{time.strftime('%Y%m%d')}-{uuid.uuid4().hex[:6].upper()}"

    with zipfile.ZipFile(io.BytesIO(zip_bytes), "r") as zf:
        file_list = [
            f for f in zf.namelist()
            if not f.startswith("__MACOSX/") and not Path(f).name.startswith(".")
            and Path(f).suffix.lower() in ALLOWED_IMG_EXTS
        ]

    if not file_list:
        raise ValueError("ZIP archive contains no valid image files (.jpg, .png, .webp, .bmp).")

    if len(file_list) > MAX_BATCH_FILES:
        file_list = file_list[:MAX_BATCH_FILES]

    init_batch_job(batch_id, len(file_list), zip_filename)

    # Launch background async worker
    asyncio.create_task(
        _run_batch_worker(batch_id, zip_bytes, file_list, upload_dir, annotated_dir, user_id)
    )

    return batch_id


async def _run_batch_worker(
    batch_id: str,
    zip_bytes: bytes,
    file_list: List[str],
    upload_dir: Path,
    annotated_dir: Path,
    user_id: Optional[str] = None
):
    """Background worker for batch processing."""
    job = BATCH_JOBS[batch_id]
    start_time = time.time()

    try:
        with zipfile.ZipFile(io.BytesIO(zip_bytes), "r") as zf:
            extracted_items = []
            for fname in file_list:
                data = zf.read(fname)
                extracted_items.append((fname, data))

        loop = asyncio.get_event_loop()

        for idx, (fname, img_bytes) in enumerate(extracted_items):
            inspection_id = f"MLX-{time.strftime('%Y%m%d')}-{uuid.uuid4().hex[:6].upper()}"
            
            # Execute in thread pool to prevent blocking event loop
            item_result = await loop.run_in_executor(
                _executor,
                process_single_batch_item,
                img_bytes,
                Path(fname).name,
                inspection_id,
                upload_dir,
                annotated_dir,
                user_id
            )

            job["inspections"].append(item_result)
            job["processed_count"] = idx + 1
            job["progress_pct"] = round(((idx + 1) / len(file_list)) * 100)

            st = item_result.get("status")
            if st == "COMPLIANT":
                job["compliant_count"] += 1
            elif st == "NON_COMPLIANT":
                job["non_compliant_count"] += 1
            else:
                job["needs_review_count"] += 1

        job["status"] = "COMPLETED"
        job["completed_at"] = time.time()
        job["duration_seconds"] = round(time.time() - start_time, 1)

    except Exception as e:
        logger.error(f"Batch processing job {batch_id} failed: {e}")
        job["status"] = "FAILED"
        job["error"] = str(e)
        job["completed_at"] = time.time()
        job["duration_seconds"] = round(time.time() - start_time, 1)


def generate_batch_excel_report(batch_id: str) -> io.BytesIO:
    """Generate professional consolidated Excel inspection report."""
    job = BATCH_JOBS.get(batch_id)
    if not job:
        raise ValueError(f"Batch job {batch_id} not found.")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Batch Inspection Summary"
    ws.views.sheetView[0].showGridLines = True

    # Palette
    NAVY = "1E293B"
    INDIGO = "4F46E5"
    LIGHT_GRAY = "F8FAFC"
    BORDER_GRAY = "E2E8F0"
    GREEN = "10B981"
    RED = "EF4444"
    AMBER = "F59E0B"

    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
    title_font = Font(name="Calibri", size=16, bold=True, color=INDIGO)
    bold_font = Font(name="Calibri", size=11, bold=True)
    regular_font = Font(name="Calibri", size=11)
    
    thin_border = Border(
        left=Side(style="thin", color=BORDER_GRAY),
        right=Side(style="thin", color=BORDER_GRAY),
        top=Side(style="thin", color=BORDER_GRAY),
        bottom=Side(style="thin", color=BORDER_GRAY),
    )

    # Title Block
    ws.merge_cells("A1:J1")
    ws["A1"] = f"MetaLex — Consolidated Legal Metrology Batch Report"
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(vertical="center")

    ws["A2"] = f"Batch Reference: {job['batch_id']}"
    ws["A2"].font = bold_font
    ws["A3"] = f"Archive Filename: {job['filename']} | Processed: {job['processed_count']}/{job['total_count']} packages"
    ws["A3"].font = regular_font
    ws["A4"] = f"Compliance: {job['compliant_count']} Compliant | {job['non_compliant_count']} Non-Compliant | {job['needs_review_count']} Needs Review"
    ws["A4"].font = bold_font

    # Headers
    headers = [
        "S.No", "Inspection ID", "Product Name", "Source File", "Status",
        "Score", "Risk Level", "MRP", "Net Quantity", "Barcode / GS1"
    ]
    
    row_idx = 6
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    # Data Rows
    for idx, item in enumerate(job.get("inspections", []), 1):
        row_idx += 1
        status = item.get("status", "PENDING")
        
        status_fill_color = "DCFCE7" if status == "COMPLIANT" else ("FEE2E2" if status == "NON_COMPLIANT" else "FEF3C7")
        status_font_color = "15803D" if status == "COMPLIANT" else ("B91C1C" if status == "NON_COMPLIANT" else "B45309")
        
        ws.cell(row=row_idx, column=1, value=idx).alignment = Alignment(horizontal="center")
        ws.cell(row=row_idx, column=2, value=item.get("inspection_id", "")).font = Font(name="Consolas", size=10)
        ws.cell(row=row_idx, column=3, value=item.get("product_name", "Unknown"))
        ws.cell(row=row_idx, column=4, value=item.get("filename", ""))
        
        status_cell = ws.cell(row=row_idx, column=5, value=status)
        status_cell.font = Font(name="Calibri", size=10, bold=True, color=status_font_color)
        status_cell.fill = PatternFill(start_color=status_fill_color, end_color=status_fill_color, fill_type="solid")
        status_cell.alignment = Alignment(horizontal="center")

        ws.cell(row=row_idx, column=6, value=f"{item.get('compliance_score', 0)}/100").alignment = Alignment(horizontal="center")
        
        risk_cell = ws.cell(row=row_idx, column=7, value=item.get("risk_level", "low").upper())
        risk_cell.alignment = Alignment(horizontal="center")
        
        ws.cell(row=row_idx, column=8, value=item.get("mrp") or "—").alignment = Alignment(horizontal="center")
        ws.cell(row=row_idx, column=9, value=item.get("net_quantity") or "—").alignment = Alignment(horizontal="center")
        ws.cell(row=row_idx, column=10, value=item.get("barcode") or "Not Detected").alignment = Alignment(horizontal="center")

        for c in range(1, 11):
            ws.cell(row=row_idx, column=c).border = thin_border

    # Auto-adjust column widths
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
